import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import date

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1E5C4A')     # teal utama (Design System Tahap 4)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
FILL_TO_EDIT = PatternFill('solid', fgColor='FFF6CC')    # kuning: sel yang WAJIB diganti user
THIN = Side(style='thin', color='D9DCE1')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
VALID_ROWS = 300  # rentang baris yang dapat divalidasi dropdown & siap diisi lanjutan

wb = openpyxl.Workbook()
wb.remove(wb.active)


def make_sheet(name, headers, col_widths, tab_color=None):
    ws = wb.create_sheet(name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 20
    for c, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width
    return ws


def write_row(ws, row_idx, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = BORDER
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
        if isinstance(v, date):
            cell.number_format = 'yyyy-mm-dd'


def add_dropdown(ws, col_letter, options_formula, first_row=2, last_row=VALID_ROWS):
    dv = DataValidation(type='list', formula1=options_formula, allow_blank=True, showErrorMessage=True)
    dv.error = 'Pilih salah satu nilai dari daftar.'
    dv.errorTitle = 'Nilai tidak valid'
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{first_row}:{col_letter}{last_row}')


# ---------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------
ws = make_sheet('CONFIG', ['key', 'value', 'keterangan'], [26, 30, 45], tab_color='9AA5B1')
config_rows = [
    ('nama_aplikasi', 'Simpan Pinjam TVRI', 'Nama aplikasi, ditampilkan di header UI'),
    ('tahun_aktif', 2026, 'Tahun operasional aktif'),
    ('folder_drive_pdf', '', 'ID folder Google Drive untuk PDF bukti transaksi — isi saat STEP P1'),
    ('folder_drive_backup', '', 'ID folder Google Drive untuk backup — isi saat STEP P1'),
    ('COUNTER_MBR_2026', 5, 'Counter terakhir ID Anggota tahun 2026 (mengikuti 5 data contoh)'),
    ('COUNTER_SP_2026', 5, 'Counter terakhir ID Simpanan tahun 2026'),
    ('COUNTER_IF_2026', 3, 'Counter terakhir ID Infaq tahun 2026'),
    ('COUNTER_PJ_2026', 3, 'Counter terakhir ID Pinjaman tahun 2026'),
    ('COUNTER_BY_2026', 10, 'Counter terakhir ID Pembayaran tahun 2026'),
    ('COUNTER_USR', 1, 'Counter terakhir ID User'),
    ('COUNTER_LOG', 0, 'Counter terakhir ID Audit Log'),
]
for i, row in enumerate(config_rows, start=2):
    write_row(ws, i, row)

# ---------------------------------------------------------------
# USERS
# ---------------------------------------------------------------
ws = make_sheet('USERS', ['user_id', 'email', 'nama', 'role', 'status', 'created_at', 'updated_at'],
                 [14, 30, 24, 12, 12, 14, 14], tab_color='9AA5B1')
write_row(ws, 2, ('USR-00001', 'GANTI-DENGAN-EMAIL-ADMIN-ANDA@tvri.go.id', 'Nama Admin Pertama',
                   'ADMIN', 'AKTIF', date(2026, 1, 1), date(2026, 1, 1)))
ws['B2'].fill = FILL_TO_EDIT
ws['B2'].comment = Comment(
    'WAJIB diganti sebelum deployment: gunakan alamat email Google (Workspace) '
    'sungguhan dari orang yang akan jadi ADMIN pertama. Auth.gs mencocokkan '
    'Session.getActiveUser().getEmail() persis dengan kolom ini.', 'Setup')
add_dropdown(ws, 'D', '"ADMIN,PETUGAS,PIMPINAN,VIEWER"')
add_dropdown(ws, 'E', '"AKTIF,NONAKTIF"')

# ---------------------------------------------------------------
# ANGGOTA
# ---------------------------------------------------------------
ws = make_sheet('ANGGOTA', ['member_id', 'nomor_anggota', 'nama', 'nik_nip', 'jenis_kelamin', 'unit',
                             'jabatan', 'no_hp', 'email', 'tanggal_bergabung', 'status', 'created_at', 'updated_at'],
                 [16, 14, 20, 14, 12, 16, 14, 16, 26, 16, 14, 14, 14], tab_color='1E5C4A')
anggota_rows = [
    ('MBR-2026-00001', 'A-001', 'Budi Santoso', '', 'L', 'Pemberitaan', 'Staff', '081234500001',
     'budi.santoso@tvri.go.id', date(2026, 1, 5), 'AKTIF', date(2026, 1, 5), date(2026, 1, 5)),
    ('MBR-2026-00002', 'A-002', 'Siti Aminah', '', 'P', 'Produksi', 'Staff', '081234500002',
     'siti.aminah@tvri.go.id', date(2026, 1, 8), 'AKTIF', date(2026, 1, 8), date(2026, 1, 8)),
    ('MBR-2026-00003', 'A-003', 'Andi Wijaya', '', 'L', 'Teknik', 'Staff', '081234500003',
     'andi.wijaya@tvri.go.id', date(2026, 2, 10), 'AKTIF', date(2026, 2, 10), date(2026, 2, 10)),
    ('MBR-2026-00004', 'A-004', 'Rina Kusuma', '', 'P', 'Keuangan', 'Staff', '081234500004',
     'rina.kusuma@tvri.go.id', date(2026, 1, 20), 'TIDAK AKTIF', date(2026, 1, 20), date(2026, 6, 1)),
    ('MBR-2026-00005', 'A-005', 'Joko Prasetyo', '', 'L', 'Pemberitaan', 'Staff', '081234500005',
     'joko.prasetyo@tvri.go.id', date(2026, 1, 12), 'AKTIF', date(2026, 1, 12), date(2026, 1, 12)),
]
for i, row in enumerate(anggota_rows, start=2):
    write_row(ws, i, row)
add_dropdown(ws, 'E', '"L,P"')
add_dropdown(ws, 'K', '"AKTIF,TIDAK AKTIF"')

# ---------------------------------------------------------------
# SIMPANAN
# ---------------------------------------------------------------
ws = make_sheet('SIMPANAN', ['transaction_id', 'member_id', 'tanggal', 'periode', 'jenis', 'nominal',
                              'metode', 'petugas', 'keterangan', 'created_at', 'status_transaksi', 'ref_koreksi'],
                 [16, 16, 12, 10, 12, 14, 12, 26, 22, 14, 14, 14], tab_color='1E5C4A')
simpanan_rows = [
    ('SP-2026-00001', 'MBR-2026-00001', date(2026, 1, 10), '2026-01', 'WAJIB', 100000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 1, 10), 'NORMAL', ''),
    ('SP-2026-00002', 'MBR-2026-00001', date(2026, 2, 10), '2026-02', 'WAJIB', 100000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 2, 10), 'NORMAL', ''),
    ('SP-2026-00003', 'MBR-2026-00002', date(2026, 1, 15), '', 'SUKARELA', 250000, 'TRANSFER',
     'admin@tvri.go.id', '', date(2026, 1, 15), 'NORMAL', ''),
    ('SP-2026-00004', 'MBR-2026-00003', date(2026, 3, 5), '2026-03', 'WAJIB', 100000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 3, 5), 'NORMAL', ''),
    ('SP-2026-00005', 'MBR-2026-00005', date(2026, 1, 20), '', 'SUKARELA', 500000, 'TRANSFER',
     'admin@tvri.go.id', '', date(2026, 1, 20), 'NORMAL', ''),
]
for i, row in enumerate(simpanan_rows, start=2):
    write_row(ws, i, row)
add_dropdown(ws, 'E', '"WAJIB,SUKARELA"')
add_dropdown(ws, 'K', '"NORMAL,VOID"')

# ---------------------------------------------------------------
# INFAQ
# ---------------------------------------------------------------
ws = make_sheet('INFAQ', ['transaction_id', 'member_id', 'tanggal', 'nominal', 'metode', 'petugas',
                           'keterangan', 'created_at', 'status_transaksi', 'ref_koreksi'],
                 [16, 16, 12, 14, 12, 26, 30, 14, 14, 14], tab_color='1E5C4A')
infaq_rows = [
    ('IF-2026-00001', 'MBR-2026-00001', date(2026, 1, 10), 50000, 'TUNAI', 'admin@tvri.go.id', '',
     date(2026, 1, 10), 'NORMAL', ''),
    ('IF-2026-00002', '', date(2026, 2, 1), 200000, 'TUNAI', 'admin@tvri.go.id',
     'Infaq umum dari donatur non-anggota', date(2026, 2, 1), 'NORMAL', ''),
    ('IF-2026-00003', 'MBR-2026-00003', date(2026, 3, 5), 25000, 'TUNAI', 'admin@tvri.go.id', '',
     date(2026, 3, 5), 'NORMAL', ''),
]
for i, row in enumerate(infaq_rows, start=2):
    write_row(ws, i, row)
add_dropdown(ws, 'I', '"NORMAL,VOID"')

# ---------------------------------------------------------------
# PINJAMAN
# ---------------------------------------------------------------
ws = make_sheet('PINJAMAN', ['loan_id', 'member_id', 'tanggal_pengajuan', 'nominal_pengajuan',
                              'tanggal_persetujuan', 'tanggal_pencairan', 'nominal_pencairan', 'tujuan',
                              'status', 'petugas', 'keterangan', 'created_at', 'updated_at',
                              'status_transaksi', 'ref_koreksi'],
                 [16, 16, 16, 16, 16, 16, 16, 22, 14, 26, 22, 14, 14, 14, 14], tab_color='1E5C4A')
pinjaman_rows = [
    ('PJ-2026-00001', 'MBR-2026-00001', date(2026, 8, 1), 10000000, date(2026, 8, 5), date(2026, 8, 10),
     10000000, 'Renovasi rumah', 'DICAIRKAN', 'admin@tvri.go.id', '', date(2026, 8, 1), date(2026, 12, 10),
     'NORMAL', ''),
    ('PJ-2026-00002', 'MBR-2026-00002', date(2026, 5, 1), 5000000, date(2026, 5, 5), date(2026, 5, 10),
     5000000, 'Biaya pendidikan anak', 'LUNAS', 'admin@tvri.go.id', '', date(2026, 5, 1), date(2026, 8, 1),
     'NORMAL', ''),
    ('PJ-2026-00003', 'MBR-2026-00003', date(2026, 8, 15), 3000000, None, None, None,
     'Kebutuhan mendesak', 'DIAJUKAN', 'admin@tvri.go.id', '', date(2026, 8, 15), date(2026, 8, 15),
     'NORMAL', ''),
]
for i, row in enumerate(pinjaman_rows, start=2):
    write_row(ws, i, row)
add_dropdown(ws, 'I', '"DIAJUKAN,DISETUJUI,DITOLAK,DICAIRKAN,LUNAS,DIBATALKAN"')
add_dropdown(ws, 'N', '"NORMAL,VOID"')

# ---------------------------------------------------------------
# PEMBAYARAN
# ---------------------------------------------------------------
ws = make_sheet('PEMBAYARAN', ['payment_id', 'loan_id', 'member_id', 'tanggal', 'nominal', 'metode',
                                'petugas', 'keterangan', 'created_at', 'status_transaksi', 'ref_koreksi'],
                 [16, 16, 16, 12, 14, 12, 26, 26, 14, 14, 14], tab_color='1E5C4A')
pembayaran_rows = [
    ('BY-2026-00001', 'PJ-2026-00001', 'MBR-2026-00001', date(2026, 9, 1), 500000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 9, 1), 'NORMAL', ''),
    ('BY-2026-00002', 'PJ-2026-00001', 'MBR-2026-00001', date(2026, 9, 15), 1500000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 9, 15), 'NORMAL', ''),
    ('BY-2026-00003', 'PJ-2026-00001', 'MBR-2026-00001', date(2026, 11, 20), 250000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 11, 20), 'NORMAL', ''),
    ('BY-2026-00004', 'PJ-2026-00001', 'MBR-2026-00001', date(2026, 12, 5), 2000000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 12, 5), 'NORMAL', ''),
    ('BY-2026-00005', 'PJ-2026-00002', 'MBR-2026-00002', date(2026, 5, 20), 1000000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 5, 20), 'NORMAL', ''),
    ('BY-2026-00006', 'PJ-2026-00002', 'MBR-2026-00002', date(2026, 6, 15), 1000000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 6, 15), 'NORMAL', ''),
    ('BY-2026-00007', 'PJ-2026-00002', 'MBR-2026-00002', date(2026, 7, 10), 1500000, 'TUNAI',
     'admin@tvri.go.id', '', date(2026, 7, 10), 'NORMAL', ''),
    ('BY-2026-00008', 'PJ-2026-00002', 'MBR-2026-00002', date(2026, 8, 1), 1500000, 'TUNAI',
     'admin@tvri.go.id', 'Melunasi pinjaman', date(2026, 8, 1), 'NORMAL', ''),
    ('BY-2026-00009', 'PJ-2026-00001', 'MBR-2026-00001', date(2026, 12, 10), 300000, 'TUNAI',
     'admin@tvri.go.id', 'Salah input nominal', date(2026, 12, 10), 'VOID', 'BY-2026-00010'),
    ('BY-2026-00010', 'PJ-2026-00001', 'MBR-2026-00001', date(2026, 12, 10), 200000, 'TUNAI',
     'admin@tvri.go.id', 'Koreksi dari BY-2026-00009', date(2026, 12, 10), 'NORMAL', ''),
]
for i, row in enumerate(pembayaran_rows, start=2):
    write_row(ws, i, row)
add_dropdown(ws, 'J', '"NORMAL,VOID"')

# ---------------------------------------------------------------
# AUDIT_LOG (sengaja kosong — lihat catatan)
# ---------------------------------------------------------------
ws = make_sheet('AUDIT_LOG', ['log_id', 'timestamp', 'user', 'action', 'module', 'record_id', 'description'],
                 [14, 18, 26, 14, 14, 16, 36], tab_color='3A4351')
ws['A2'].comment = Comment(
    'Sengaja dikosongkan: data contoh di sheet lain diisi langsung ke spreadsheet '
    'untuk keperluan uji coba, bukan lewat aplikasi — jadi tidak punya jejak audit. '
    'Transaksi yang dibuat lewat aplikasi nanti akan otomatis tercatat di sini.', 'Catatan')

wb.save('/home/claude/simpan-pinjam-gas/database/Database_Simpan_Pinjam.xlsx')
print('saved')
